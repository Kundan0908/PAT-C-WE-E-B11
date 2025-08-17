import datetime
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.ie.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pytest
import openpyxl

def load_data():
    excel_path = 'C:/Users/Kundan_Kumar/PycharmProjects/Guvi-PAT-C-WE-E-B11/Data_Driven/test_data.xlsx'
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['test_data']
    max_row_ = sheet.max_row #2
    max_col_ = sheet.max_column #2
    print(f"Maximum row count {max_row_} and maximimum column count is {max_col_}")
    data_list = [] # []
    for row in range(2,max_row_+1):
        username = sheet.cell(row=row,column=1).value
        password = sheet.cell(row=row,column=2).value
        data_list.append((username,password))
    return data_list


@pytest.mark.parametrize('username,password',load_data())
def test_total_balance_remaining(username,password):
    try:
        # Invoking Browser and maximising window
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.get("https://parabank.parasoft.com/parabank/index.htm?ConnType=JDBC")
        driver.maximize_window()

        # logging in to account
        driver.find_element(By.NAME,'username').send_keys(username)
        driver.find_element(By.NAME,'password').send_keys(password)
        driver.find_element(By.XPATH, "//*[@type='submit']").click()
        time.sleep(4)
        driver.find_element(By.XPATH,"//*[text()='Log Out']").click()

    finally:
        driver.quit()

# def test_register_user():
#     try:
#         driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
#         driver.get("https://parabank.parasoft.com/parabank/index.htm?ConnType=JDBC")
#         driver.find_element(By.XPATH,'//a[text()="Register"]').click()
#         driver.find_element(By.NAME,'customer.firstName').send_keys('testuser')
#         driver.find_element(By.NAME, 'customer.lastName').send_keys('testuser')
#         driver.find_element(By.NAME, 'customer.address.street').send_keys('430,street road')
#         driver.find_element(By.NAME, 'customer.address.city').send_keys('Bangalore')
#         driver.find_element(By.NAME, 'customer.address.state').send_keys('karnataka')
#         driver.find_element(By.NAME, 'customer.address.zipCode').send_keys('1234238')
#         driver.find_element(By.NAME, 'customer.ssn').send_keys('1234238')
#         driver.find_element(By.NAME, 'customer.username').send_keys('testuser12')
#         driver.find_element(By.NAME, 'customer.password').send_keys('testuser')
#         driver.find_element(By.NAME, 'repeatedPassword').send_keys('testuser')
#         driver.find_element(By.XPATH,'//input[@value="Register"]').click()
#         actual_output = driver.find_element(By.XPATH,"//*[@id='rightPanel']//child::p").text
#         expected_output = 'Your account was created successfully. You are now logged in.'
#         driver.save_screenshot('image.png')
#         assert  expected_output == actual_output, "Account not created due to some problem"
#
#     except:
#         raise Exception
#
#     finally:
#         driver.quit()