import openpyxl

def read_data():
    workbk = openpyxl.load_workbook('test_data.xlsx')
    wrksheet = workbk['test_data']
    data = []
    for each_row in range(2,wrksheet.max_row+1):
        first_name = wrksheet.cell(each_row,1).value
        last_name = wrksheet.cell(each_row,2).value
        address = wrksheet.cell(each_row,3).value
        city = wrksheet.cell(each_row,4).value
        state = wrksheet.cell(each_row,5).value
        phnumber = wrksheet.cell(each_row,6).value
        ssn = wrksheet.cell(each_row,7).value
        username = wrksheet.cell(each_row,8).value
        password = wrksheet.cell(each_row,9).value
        rptdpassword = wrksheet.cell(each_row,10).value
        user_data = (first_name,last_name,address,city,state,phnumber,ssn,username,password,rptdpassword) # (a,b,c),(d,e,f)
        data.append(user_data) # [(a,b,c),(d,e,f)]
    workbk.close()
    return data