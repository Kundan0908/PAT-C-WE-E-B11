import openpyxl
from Tasks.Task_14.POM.Login import LoginPage
import pytest

# Data-driven test using Excel
class ExcelDataProvider:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_test_data(self):
        test_data = []
        for row in range(2, self.sheet.max_row + 1):  # Skip header row
            test_id = self.sheet.cell(row=row, column=1).value
            username = self.sheet.cell(row=row, column=2).value
            password = self.sheet.cell(row=row, column=3).value
            test_data.append((test_id, username, password)) # [(test_id,username,password)]
        return test_data

    def write_result(self, test_id, result):
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == test_id:
                self.sheet.cell(row=row, column=4).value = result
                break
        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()

# Test case
def test_login_with_excel_data(driver):
    # Initialize data provider
    data_provider = ExcelDataProvider("test_data.xlsx")
    test_cases = data_provider.get_test_data() #[(test_id,username,password)]

    # Initialize page object
    login_page = LoginPage(driver)

    for test_id, username, password in test_cases:
        # Navigate to URL
        login_page.navigate_to_url("https://www.saucedemo.com/")

        # Perform login
        login_page.enter_username(username)
        login_page.enter_password(password)
        login_page.click_login()

        # Check result and write to Excel
        if login_page.is_login_successful():
            data_provider.write_result(test_id, "PASS")
        elif login_page.is_login_failed():
            data_provider.write_result(test_id, "FAIL")
        else:
            data_provider.write_result(test_id, "ERROR")

    data_provider.close()

